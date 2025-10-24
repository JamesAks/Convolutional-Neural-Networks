import openpyxl as exl
import time

class EvalLogger:

    # Logs the evaluation of the models into a Ecxel Sheet

    def __init__(self, filename: str ="ModelEvaluation.xlsx", mode: str = ""):

        self.filename = filename
        self.mode = mode
        self.wb = exl.load_workbook(filename)

        self.sheetnames = self.wb.sheetnames
        self.sheets = {}

        for sheet in self.wb:

            self.sheets[sheet.title] = sheet

        self.start_time = time.time()

        # Remove default sheet if no models are defined yet

        # Create a sheet for each model


    def createSheet(self, model_name):
        
        # Creates Excel sheet using models name

        ws = self.wb.create_sheet(title = model_name)

        if self.mode == "CV":

            ws = self.wb.create_sheet(title = f"{model_name} - CV")
            ws.append(["Test Fold", "Epoch", "Batch", "Loss", "Accuracy", "Runtime(s)"])
            self.sheets[model_name] = ws
            
        else:

            ws = self.wb.create_sheet(title = model_name)
            ws.append(["Epoch", "Batch", "Loss", "Accuracy", "Runtime(s)"])
            self.sheets[model_name] = ws

        self.save()

    
    def log(self, model_name, fold, epoch, batch, loss, acc):

        elapsed = time.time() - self.start_time
        
        if self.mode == "CV" :

            ws = self.sheets[f"{model_name} - CV"]
            ws.append([fold, epoch, batch, loss, acc, elapsed])

        else:

            ws = self.sheets[model_name]
            ws.append([epoch, batch, loss, acc, elapsed])

        self.save()


    def testLog(self, model_name, acc, index: int = 1):

        # logging of the testing results of models

        if self.mode == "CV":

            ws = self.sheets[f"{model_name} - CV"]
            ws.cell(row = index, column = 7).value = f"Fold {index} Test Result:"
            ws.cell(row = index, column = 8).value = acc

        else:

            ws = self.sheets[model_name]
            ws.cell(row = index, column = 7).value = "Test Result:"
            ws.cell(row = index, column = 8).value = acc

        self.save()


    def save(self):

        self.wb.save(self.filename)


    def checkWorksheet(self, model_name):

        # Checks for Worksheet and replaces it with a new fresh sheet
        
        if model_name in self.sheetnames:

            for row in self.sheets[model_name].iter_rows():
                for cell in row:

                    cell.value = None

            self.wb.save(model_name)

            self.createSheet(model_name)




     