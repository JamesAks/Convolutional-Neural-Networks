import numpy as np
import cupy as cp
from DataLoader import DataLoader
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
# from Instantiate import instantiateModels
from Losses import logLoss, logLossPrime
import openpyxl
import time


class TrainingLogger:
    def __init__(self, filename: str ="ModelEvaluation.xlsx"):

        self.filename = filename
        self.wb = load_workbook(filename)

        self.sheets = {}
        self.start_time = time.time()

        # Remove default sheet if no models are defined yet

        # Create a sheet for each model

    def create_sheet(self, model_name):

        ws = self.wb.create_sheet(title = model_name)
        ws.append(["Fold", "Epoch", "Batch", "Loss", "Accuracy", "Runtime (s)"])
        self.sheets[model_name] = ws

    def log(self, model_name, fold, epoch, batch, loss, acc):
        elapsed = time.time() - self.start_time
        ws = self.sheets[model_name]
        ws.append([fold, epoch, batch, loss, acc, elapsed])

    def save(self):
        self.wb.save(self.filename)



def predict(network: list, inputs: np.ndarray):

    output = inputs

    for layer in network:

        output = layer.forward(output)
        
    return output


def train(model_name: str, network: list, loss, lossPrime, dataLoader: DataLoader, learningRate : float = 0.01, epochs: int = 100, ):

    trainLogger = TrainingLogger()
    trainLogger.create_sheet(model_name)

    for e in range(epochs):

        loss_sum = 0
        batch = 0

        for x,y in dataLoader:

            outputs = predict(network, x)

            print(f"Final output size: {outputs.size}" )
            print(f"Input Type: {type(x)} and Value Type: {type(y)}")


            loss_sum += float(loss(outputs, y))

            error_gradient = lossPrime(outputs, y)

            acc = float(findAccuracy(outputs, y))

            for layer in reversed(network):

                output = layer.backward(error_gradient, learningRate)
                layer.update()

                error_gradient = output

            print(loss_sum)

            trainLogger.log(model_name,1,(e + 1), (batch + 1), loss_sum, acc)
            trainLogger.save()
        
        print(f"Epoch: {e + 1} --- Loss: {loss_sum}")


def findAccuracy(outputs, y) -> float:

    if type(outputs) == np.ndarray:
        module = np

    else:
        module = cp

    predicts = module.argmax(outputs, axis = 1)
    correct = (module.argmax(y, axis = 1))


    return module.mean(predicts == correct)






# def test(network,x_test, y_test):

#     correct = 0

#     for x,y in zip(x_test, y_test):


#         if output == y:

#             correct += 1

#     print(f"Top-1 Accuracy: {correct/x_test.size}")


# def cross_validate(models: list, model_names: list, input_images, labels,logger: TrainingLogger, loss, lossPrime, k: int=5, epochs: int=10, batch_size: int=32, learning_rate: float = 0.01):

#     # Cross Validation is for comparing the models and done after the initial training
    
#     data_size = input_images.shape[0]
#     indices = np.arange(data_size)
#     np.random.shuffle(indices)
    
#     fold_sizes = np.full(k, data_size // k, dtype=int)

#     fold_sizes[:data_size % k] += 1

#     current = 0
    
#     for fold_idx, fold_size in enumerate(fold_sizes):

#         start, stop = current, current + fold_size

#         val_idx = indices[start:stop]
#         train_idx = np.concatenate([indices[:start], indices[stop:]])
#         current = stop
        
#         X_train, y_train = input_images[train_idx], labels[train_idx]
#         X_test, y_test = input_images[val_idx], labels[val_idx]
        
#         print(f"\n--- Fold {fold_idx+1}/{k} ---")
        
#         for model, name in zip(models, model_names):

#             print(f"Training model: {name}")
            
#             for epoch in range(epochs):

#                 # Mini-batch training
#                 for batch_start in range(0, len(X_train), batch_size):

#                     batch_end = batch_start + batch_size
#                     batch_inputs = X_train[batch_start:batch_end]
#                     batch_labels = y_train[batch_start:batch_end]
                    
#                     # Forward
#                     outputs = model.forward(batch_inputs)
                    
#                     # Compute loss

#                     loss = loss(outputs, batch_inputs)
                    
#                     # Backward
#                     error_gradient = lossPrime(outputs, batch_inputs)
                    
#                     model.backward(error_gradient)
                    
#                     # Update parameters
#                     model.update(learning_rate)
                    
#                     # Compute accuracy

            
                    

#             test_outputs = model.forward(X_test)
#             test_loss = lossPrime(test_outputs, y_test)
#             test_acc = findAccuracy(test_outputs, y_test)

#             logger.log(name, fold_idx + 1, epochs, 'val', float(test_loss), float(test_acc))

#             print(f"Validation after training: Loss: {test_loss:.4f}, Acc: {test_acc:.4f}")
        
#             #print(f"Epoch {epoch+1}/{epochs}, Loss: {loss:.4f}, Accuracy: {acc:.4f}")

#     return()

image_path = "C:/Users/james/Documents/Masters/Dissertation/Preprocessed_Train/Train_images"
train_exp_path = "C:/Users/james/Documents/Masters/Dissertation/Preprocessed_Train/Train_expected"

data = DataLoader(image_path, train_exp_path, 10000, targetSize=(32,32), numb_samples= 10000) 
gpu_data = DataLoader(image_path, train_exp_path, 10000, targetSize=(244,244), numb_samples= 10000, gpu = True)



# def Run(models:list, model_names: list, data: DataLoader):

#     crossLogger = TrainingLogger()

#     for x,y in data:

#         cross_validate(models, model_names, x, y, crossLogger, logLoss, logLossPrime, 5, 20, 32, 0.01)


# models, model_names = instantiateModels("Normal")

# Run(models,model_names,data)













        


        
    





